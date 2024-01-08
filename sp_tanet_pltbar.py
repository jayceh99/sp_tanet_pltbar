import openpyxl
import csv
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
class c_sp_pltbar:
    def __init__(self , csv_filename) -> None:
        self.csv_filename = csv_filename
        self.round_ = {'v4_Download':2 , 'v4_Upload':4 , 'v6_Download':3 , 'v6_Upload':5}

    def f_get_all_value(self):
        csv_data = []
        f = open(self.csv_filename , 'r' , encoding='utf-8')
        for i in csv.reader(f) : 
            csv_data.append(i)
        workbook = openpyxl.workbook.Workbook()
        worksheet = workbook.active
        for row in csv_data:
            worksheet.append(row)
        self.sheet = workbook.worksheets[0]
        del csv_data , f , workbook ,row

    def f_plt6m_bar(self):
        for k in self.round_:
            range5_ = 0
            range5_7 = 0
            range7_9 = 0
            range9_ = 0
            #<5 5~7 7~9 >9
            for i in range(6 , self.sheet.max_row+1):
                value = float(self.sheet.cell(row = i , column = self.round_[k]).value)
                if value < 5 :
                    range5_ += 1
                elif value >= 5 and value < 7 :
                    range5_7 += 1
                elif value >= 7 and value < 9 :
                    range7_9 += 1
                elif value >= 9 :
                    range9_ += 1
            x = [ 1 , 2 , 3 , 4 ]        
            label = [ '<5' , '5~7' , '7~9' , '>9' ]     
            h = [ range5_ , range5_7 , range7_9 , range9_ ]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            plt.title(k,fontsize=24)
            y_major_locator=MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.get_current_fig_manager().window.state('zoomed')
            #plt.show()
            plt.savefig(fname = "C:\\Users\\jayce\\Desktop\\test_data\\6m"+k+".png" )
            plt.cla()

    def f_plt10m_bar(self):
        for k in self.round_:
            range5_ = 0
            range5_7 = 0
            range7_9 = 0
            range9_11 = 0
            range11_ = 0
            #<5 5~7 7~9 >9
            for i in range(6 , self.sheet.max_row+1):
                value = float(self.sheet.cell(row = i , column = self.round_[k]).value)
                if value < 5 :
                    range5_ += 1
                elif value >= 5 and value < 7 :
                    range5_7 += 1
                elif value >= 7 and value < 9 :
                    range7_9 += 1
                elif value >= 9 and value < 11 :
                    range9_11 += 1
                elif value >= 11 :
                    range11_ += 1
            x = [ 1 , 2 , 3 , 4 , 5 ]        
            label = [ '<5' , '5~7' , '7~9' , '9~11' , '>11' ]     
            h = [ range5_ , range5_7 , range7_9 , range9_11 , range11_]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            plt.title(k,fontsize=24)
            y_major_locator=MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.get_current_fig_manager().window.state('zoomed')
            #plt.show()
            plt.savefig(fname = "C:\\Users\\jayce\\Desktop\\test_data\\10m"+k+".png" )
            plt.cla()

    def f_plt_bar(self):
        for k in self.round_:
            range0_2 = 0
            range2_4 = 0
            range4_6 = 0
            range6_8 = 0
            range8_10 = 0
            range10_12 = 0
            range12_14 = 0
            range14_16 = 0
            range16_18 = 0
            range18_ = 0
            for i in range(6 , self.sheet.max_row+1):
                try:
                    value = float(self.sheet.cell(row = i , column = self.round_[k]).value)
                    if value >= 0 and value < 2 :
                        range0_2 += 1
                    elif value >= 2 and value < 4 :
                        range2_4 += 1
                    elif value >= 4 and value < 6 :
                        range4_6 += 1
                    elif value >= 6 and value < 8 :
                        range6_8 += 1
                    elif value >= 8 and value < 10 :
                        range8_10 += 1
                    elif value >= 10 and value < 12 :
                        range10_12 += 1
                    elif value >= 12 and value < 14 :
                        range12_14 += 1    
                    elif value >= 14 and value < 16 :
                        range14_16 += 1   
                    elif value >= 16 and value < 18 :
                        range16_18 += 1  
                    elif value >= 18 :
                        range18_ += 1
                except:
                    continue
            x = [1,2,3,4,5,6,7,8,9,10]        
            label = ['0~2','2~4','4~6','6~8','8~10','10~12','12~14','14~16','16~18','>18']     
            h = [ range0_2 , range2_4 , range4_6 , range6_8 , range8_10 , range10_12 , range12_14 , range14_16 , range16_18 , range18_]   
            fig = plt.figure()
            fig.set_size_inches(12,9)
            plt.title(k,fontsize=24)
            y_major_locator=MultipleLocator(2)
            ax=plt.gca()
            ax.yaxis.set_major_locator(y_major_locator)
            plt.bar(x,h,tick_label=label,width=0.5)  
            plt.get_current_fig_manager().window.state('zoomed')
            #plt.show()
            plt.savefig(fname = "C:\\Users\\jayce\\Desktop\\test_data\\"+k+".png" )
            plt.cla()

def main ():
    csv_filename = (r'C:\Users\jayce\Desktop\test_data\Group test results 2023-12-22 13_36_08.csv')
    sp_pltbar = c_sp_pltbar(csv_filename)
    sp_pltbar.f_get_all_value()
    sp_pltbar.f_plt_bar()

if __name__ == "__main__" :
    main()